# from playwright.sync_api import sync_playwright
# from openpyxl import load_workbook
#
# # Load Excel
# wb = load_workbook("Data.xlsx")
# sheet = wb.active
#
# with sync_playwright() as p:
#     browser = p.chromium.launch(headless=False)
#     page = browser.new_page()
#
#     # Loop through Excel rows
#     for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
#         username, password = row
#
#         print(f"Running test for: {username}")
#
#         page.goto("https://www.saucedemo.com/")
#
#         page.fill("#user-name", username)
#         page.fill("#password", password)
#         page.click("#login-button")
#
#         # Validation (success or error)
#         if "inventory" in page.url:
#             print("✅ Login Success")
#             sheet[f"C{i}"] = "Pass"
#
#             # logout for next iteration
#             page.click("#react-burger-menu-btn")
#             page.click("#logout_sidebar_link")
#
#         else:
#             print("❌ Login Failed")
#             sheet[f"C{i}"] = "Fail"
#
#     # Save results back to Excel
#     wb.save("Data.xlsx")
#
#     browser.close()

############################################################################################
#another way

# from playwright.sync_api import sync_playwright
# from openpyxl import load_workbook
#
# wb = load_workbook("Data.xlsx")
# sheet = wb.active
#
# with sync_playwright() as p:
#     browser = p.chromium.launch(headless=False)
#     page = browser.new_page()
#
#     # loop using row numbers
#     for i in range(2, sheet.max_row + 1):
#         username = sheet.cell(row=i, column=1).value
#         password = sheet.cell(row=i, column=2).value
#
#         print(f"Running test for: {username}")
#
#         page.goto("https://www.saucedemo.com/")
#
#         page.fill("#user-name", username)
#         page.fill("#password", password)
#         page.click("#login-button")
#
#         if "inventory" in page.url:
#             print("✅ Pass")
#             sheet.cell(row=i, column=3).value = "Pass"
#
#             page.click("#react-burger-menu-btn")
#             page.click("#logout_sidebar_link")
#         else:
#             print("❌ Fail")
#             sheet.cell(row=i, column=3).value = "Fail"
#
#     wb.save("Data.xlsx")
#     browser.close()


#############################################################################################################
#async


import asyncio
from playwright.async_api import async_playwright
from openpyxl import load_workbook


async def run_test():
    # Load Excel
    wb = load_workbook("Data.xlsx")
    sheet = wb.active

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context()
        page = await context.new_page()

        # Loop using row index (no enumerate)
        for i in range(2, sheet.max_row + 1):
            username = sheet.cell(row=i, column=1).value
            password = sheet.cell(row=i, column=2).value

            print(f"Running test for: {username}")

            await page.goto("https://www.saucedemo.com/")

            await page.fill("#user-name", username)
            await page.fill("#password", password)
            await page.click("#login-button")

            # Validate result
            if "inventory" in page.url:
                print("✅ Pass")
                sheet.cell(row=i, column=3).value = "Pass"

                # Logout for next iteration
                await page.click("#react-burger-menu-btn")
                await page.click("#logout_sidebar_link")
            else:
                print("❌ Fail")
                sheet.cell(row=i, column=3).value = "Fail"

        # Save results
        wb.save("Data.xlsx")

        await browser.close()


# Run async function
asyncio.run(run_test())